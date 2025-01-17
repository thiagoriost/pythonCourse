import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

def generar_archivo_excel(nombre_archivo):
    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Pago"

    # Encabezados de entrada
    ws["A1"] = "Monto Prestado"
    ws["B1"] = "Tasa de Interés Anual (%)"
    ws["C1"] = "Plazo (meses)"

    # Valores de ejemplo
    ws["A2"] = 1000000  # Ejemplo: $1,000,000
    ws["B2"] = 12        # Ejemplo: 12%
    ws["C2"] = 12        # Ejemplo: 12 meses

    # Encabezados de salida
    ws["A4"] = "N° Cuota"
    ws["B4"] = "Fecha de Vencimiento"
    ws["C4"] = "Interés"
    ws["D4"] = "Amortización"
    ws["E4"] = "Cuota"
    ws["F4"] = "Saldo Pendiente"

    # Formato de encabezados
    for col in "ABCDEF":
        ws[f"{col}4"].alignment = Alignment(horizontal="center", vertical="center")

    # Cálculos dinámicos
    monto = ws["A2"].value
    tasa_anual = ws["B2"].value / 100
    plazo = ws["C2"].value

    tasa_mensual = tasa_anual / 12
    cuota = monto * (tasa_mensual * (1 + tasa_mensual) ** plazo) / ((1 + tasa_mensual) ** plazo - 1)

    saldo = monto
    fecha_inicio = datetime.today()

    for i in range(1, plazo + 1):
        interes = saldo * tasa_mensual
        amortizacion = cuota - interes
        saldo -= amortizacion
        fecha_vencimiento = fecha_inicio + timedelta(days=30 * i)

        ws.append([
            i,
            fecha_vencimiento.strftime("%d/%m/%Y"),
            round(interes, 2),
            round(amortizacion, 2),
            round(cuota, 2),
            round(saldo, 2)
        ])

    # Guardar el archivo Excel
    wb.save(nombre_archivo)

# Generar el archivo Excel
generar_archivo_excel("control_credito.xlsx")
